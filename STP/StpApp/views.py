from xhtml2pdf import pisa
from django.http import HttpResponse
from io import BytesIO
from django.shortcuts import render, HttpResponse, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .models import RegForm, TUpload
import math
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
import os
import smtplib
from email.message import EmailMessage
from django.template.loader import get_template
from django.views.generic import View
from django.core.files.storage import FileSystemStorage


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


class GeneratePDF(View):
    def get(self, request, *args, **kwargs):
        template = get_template('STPApp/cert.html')
        context = {
            "results": "Abhinav Raturi",
            "course": "Internship in STP",
            "grades": "A+"
        }
        html = template.render(context)
        pdf = render_to_pdf('STPApp/cert.html', context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "results_%s.pdf" % ("12341231")
            content = "inline; filename='%s'" % (filename)
            download = request.GET.get("download")
            if download:
                content = "attachment; filename='%s'" % (filename)
            response['Content-Disposition'] = content
            return response
        return HttpResponse("Not found")


def home(request):

    return render(request, 'STPApp/home.html')


def register(request):
    if request.method == 'POST':
        name = request.POST['name']
        dob = request.POST['dob']
        gender = request.POST['gender']
        address = request.POST['address']
        mobile = request.POST['mobile']
        email = request.POST['mailadd']
        aadhar = int(request.POST['aadhar'])
        fname = request.POST['father']
        foccupation = request.POST['occupation']
        institutename = request.POST['institute']
        course = request.POST['course']
        presentyear = request.POST['pyear']
        percentage12 = request.POST['std12']
        cgpa = request.POST['cgpa']
        photograph = request.FILES['photo']
        marksheet12 = request.FILES['marksheet12']
        lmarksheet = request.FILES['dmc']
        bletter = request.FILES['certificate']
        details = RegForm(name=name, dob=dob, gender=gender, address=address, mobile=mobile, email=email, aadhar=aadhar, fname=fname, foccupation=foccupation, institutename=institutename, course=course, presentyear=presentyear, percentage12=percentage12,
                          cgpa=cgpa, photograph=photograph, marksheet12=marksheet12, lmarksheet=lmarksheet, bletter=bletter)
        details.save()
        print(name)
        print('we are using post request')
        all_xls_sheet()
        selected_xls_sheet()
        rejected_xls_sheet()
        selected_xls_sheet()
        messages.success(request, "Form submitted Successfully")
    return render(request, 'STPApp/registration.html')


def handleofficerLogin(request):
    if request.method == 'POST':
        officerusername = request.POST['officerusername']
        officerpass = request.POST['officerpass']
        user = authenticate(username=officerusername, password=officerpass)
        if user is not None:
            usernameval1 = str(user.username)
            name = usernameval1[2:5]
            if name == "STP":
                login(request, user)
                return redirect("officer")
            else:
                messages.error(request, "Invalid credentials")
                return redirect("/")
        else:
            messages.error(request, "Invalid credentials")
            return redirect("/")
    return HttpResponse('404 - Not found')


def handleofficerLogout(request):

    logout(request)
    messages.success(request, "Successfully logout")
    return redirect("/")


@login_required(login_url='officerlogin')
def handelofficer(request):
    RegForms = RegForm.objects.all()
    n = len(RegForms)
    params = {'allforms': RegForms, 'range': range(n)}
    return render(request, 'STPApp/officer.html', params)


@login_required(login_url='officerlogin')
def deleterecent(request):
    if request.method == 'POST':
        id12 = request.POST['dvalue']
        d = RegForm.objects.get(sno=id12)
        d.counter = 1
        d.save()
        selected_xls_sheet()
        all_xls_sheet()
        rejected_xls_sheet()
        selected_xls_sheet()
        return redirect("officer")


@login_required(login_url='officerlogin')
def acceptrecent(request):
    if request.method == 'POST':
        id12 = request.POST['dvalue1']
        d = RegForm.objects.get(sno=id12)
        d.counter = 2
        d.save()
        selected_xls_sheet()
        all_xls_sheet()
        rejected_xls_sheet()
        selected_xls_sheet()
        return redirect("officer")


def selected_xls_sheet():
    RegForms = RegForm.objects.all()
    total_no = len(RegForms)
    c = 2
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 50
    ws1.column_dimensions['H'].width = 20
    ws1.column_dimensions['I'].width = 20
    ws1.column_dimensions['J'].width = 25
    ws1.column_dimensions['K'].width = 25
    ws1.column_dimensions['L'].width = 15
    ws1.column_dimensions['M'].width = 15
    ws1.column_dimensions['N'].width = 15
    ws1.column_dimensions['O'].width = 20
    ws1.column_dimensions['P'].width = 12
    ws1.column_dimensions['Q'].width = 12
    ws1.column_dimensions['R'].width = 12
    ws1.column_dimensions['S'].width = 20
    ws1.column_dimensions['T'].width = 20
    (ws1.cell(row=1, column=1, value="Name")).font = Font(bold=True)
    (ws1.cell(row=1, column=2, value="DOB")).font = Font(bold=True)
    (ws1.cell(row=1, column=3, value="Gender")).font = Font(bold=True)
    (ws1.cell(row=1, column=4, value="E-mail")).font = Font(bold=True)
    (ws1.cell(row=1, column=5, value="Adhaar Number")).font = Font(bold=True)
    (ws1.cell(row=1, column=6, value="Mobile")).font = Font(bold=True)
    (ws1.cell(row=1, column=7, value="Address")).font = Font(bold=True)
    (ws1.cell(row=1, column=8, value="Father/Mother Name")).font = Font(bold=True)
    (ws1.cell(row=1, column=9, value="Occupation")).font = Font(bold=True)
    (ws1.cell(row=1, column=10, value="Institute Name")).font = Font(bold=True)
    (ws1.cell(row=1, column=11, value="Course")).font = Font(bold=True)
    (ws1.cell(row=1, column=12, value="Present Year")).font = Font(bold=True)
    (ws1.cell(row=1, column=13, value="12th Percentage")).font = Font(bold=True)
    (ws1.cell(row=1, column=14, value="CGPA")).font = Font(bold=True)

    for i in range(total_no):
        name = RegForms[i].name
        dob = str(RegForms[i].dob)
        gender = RegForms[i].gender
        address = RegForms[i].address
        mobile = RegForms[i].mobile
        email = RegForms[i].email
        aadhar = str(RegForms[i].aadhar)
        fname = RegForms[i].fname
        foccupation = RegForms[i].foccupation
        institutename = RegForms[i].institutename
        course = RegForms[i].course
        presentyear = RegForms[i].presentyear
        percentage12 = RegForms[i].percentage12
        cgpa = RegForms[i].cgpa

        if RegForms[i].counter == 2:
            ws1.cell(row=c, column=1, value=name)
            ws1.cell(row=c, column=2, value=dob)
            ws1.cell(row=c, column=3, value=gender)
            ws1.cell(row=c, column=4, value=email)
            ws1.cell(row=c, column=5, value=aadhar)
            ws1.cell(row=c, column=6, value=mobile)
            ws1.cell(row=c, column=7, value=address)
            ws1.cell(row=c, column=8, value=fname)
            ws1.cell(row=c, column=9, value=foccupation)
            ws1.cell(row=c, column=10, value=institutename)
            ws1.cell(row=c, column=11, value=course)
            ws1.cell(row=c, column=12, value=presentyear)
            ws1.cell(row=c, column=13, value=percentage12)
            ws1.cell(row=c, column=14, value=cgpa)

            c = c+1
        wb1.save("media/xl_sheets1/selected_one_xl_file.xlsx")


def rejected_xls_sheet():
    RegForms = RegForm.objects.all()
    total_no = len(RegForms)
    c = 2
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 12
    ws.column_dimensions['Q'].width = 12
    ws.column_dimensions['R'].width = 12
    ws.column_dimensions['S'].width = 20
    ws.column_dimensions['T'].width = 20
    (ws.cell(row=1, column=1, value="Name")).font = Font(bold=True)
    (ws.cell(row=1, column=2, value="DOB")).font = Font(bold=True)
    (ws.cell(row=1, column=3, value="Gender")).font = Font(bold=True)
    (ws.cell(row=1, column=4, value="E-mail")).font = Font(bold=True)
    (ws.cell(row=1, column=5, value="Adhaar Number")).font = Font(bold=True)
    (ws.cell(row=1, column=6, value="Mobile")).font = Font(bold=True)
    (ws.cell(row=1, column=7, value="Address")).font = Font(bold=True)
    (ws.cell(row=1, column=8, value="Father/Mother Name")).font = Font(bold=True)
    (ws.cell(row=1, column=9, value="Occupation")).font = Font(bold=True)
    (ws.cell(row=1, column=10, value="Institute Name")).font = Font(bold=True)
    (ws.cell(row=1, column=11, value="Course")).font = Font(bold=True)
    (ws.cell(row=1, column=12, value="Present Year")).font = Font(bold=True)
    (ws.cell(row=1, column=13, value="12th Percentage")).font = Font(bold=True)
    (ws.cell(row=1, column=14, value="CGPA")).font = Font(bold=True)

    for i in range(total_no):
        name = RegForms[i].name
        dob = str(RegForms[i].dob)
        gender = RegForms[i].gender
        address = RegForms[i].address
        mobile = RegForms[i].mobile
        email = RegForms[i].email
        aadhar = str(RegForms[i].aadhar)
        fname = RegForms[i].fname
        foccupation = RegForms[i].foccupation
        institutename = RegForms[i].institutename
        course = RegForms[i].course
        presentyear = RegForms[i].presentyear
        percentage12 = RegForms[i].percentage12
        cgpa = RegForms[i].cgpa

        if RegForms[i].counter == 1:
            ws.cell(row=c, column=1, value=name)
            ws.cell(row=c, column=2, value=dob)
            ws.cell(row=c, column=3, value=gender)
            ws.cell(row=c, column=4, value=email)
            ws.cell(row=c, column=5, value=aadhar)
            ws.cell(row=c, column=6, value=mobile)
            ws.cell(row=c, column=7, value=address)
            ws.cell(row=c, column=8, value=fname)
            ws.cell(row=c, column=9, value=foccupation)
            ws.cell(row=c, column=10, value=institutename)
            ws.cell(row=c, column=11, value=course)
            ws.cell(row=c, column=12, value=presentyear)
            ws.cell(row=c, column=13, value=percentage12)
            ws.cell(row=c, column=14, value=cgpa)

            c = c+1
    wb.save("media/xl_sheets1/rejected_one_xl_file.xlsx")


def all_xls_sheet():
    RegForms = RegForm.objects.all()
    total_no = len(RegForms)
    c = 2
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 12
    ws.column_dimensions['R'].width = 12
    ws.column_dimensions['S'].width = 12
    ws.column_dimensions['T'].width = 20
    ws.column_dimensions['U'].width = 20
    (ws.cell(row=1, column=1, value="Name")).font = Font(bold=True)
    (ws.cell(row=1, column=2, value="Status")).font = Font(bold=True)
    (ws.cell(row=1, column=3, value="DOB")).font = Font(bold=True)
    (ws.cell(row=1, column=4, value="Gender")).font = Font(bold=True)
    (ws.cell(row=1, column=5, value="E-mail")).font = Font(bold=True)
    (ws.cell(row=1, column=6, value="Adhaar Number")).font = Font(bold=True)
    (ws.cell(row=1, column=7, value="Mobile")).font = Font(bold=True)
    (ws.cell(row=1, column=8, value="Address")).font = Font(bold=True)
    (ws.cell(row=1, column=9, value="Father/Mother Name")).font = Font(bold=True)
    (ws.cell(row=1, column=10, value="Occupation")).font = Font(bold=True)
    (ws.cell(row=1, column=11, value="Institute Name")).font = Font(bold=True)
    (ws.cell(row=1, column=12, value="Course")).font = Font(bold=True)
    (ws.cell(row=1, column=13, value="Present Year")).font = Font(bold=True)
    (ws.cell(row=1, column=14, value="12th Percentage")).font = Font(bold=True)
    (ws.cell(row=1, column=15, value="CGPA")).font = Font(bold=True)

    for i in range(total_no):
        if RegForms[i].counter == 1 or RegForms[i].counter == 2:
            if RegForms[i].counter == 1:
                status = "ACCEPTED"
            else:
                status = "REJECTED"
            name = RegForms[i].name
            dob = str(RegForms[i].dob)
            gender = RegForms[i].gender
            address = RegForms[i].address
            mobile = RegForms[i].mobile
            email = RegForms[i].email
            aadhar = str(RegForms[i].aadhar)
            fname = RegForms[i].fname
            foccupation = RegForms[i].foccupation
            institutename = RegForms[i].institutename
            course = RegForms[i].course
            presentyear = RegForms[i].presentyear
            percentage12 = RegForms[i].percentage12
            cgpa = RegForms[i].cgpa

            ws.cell(row=c, column=1, value=name)
            ws.cell(row=c, column=2, value=status)
            ws.cell(row=c, column=3, value=dob)
            ws.cell(row=c, column=4, value=gender)
            ws.cell(row=c, column=5, value=email)
            ws.cell(row=c, column=6, value=aadhar)
            ws.cell(row=c, column=7, value=mobile)
            ws.cell(row=c, column=8, value=address)
            ws.cell(row=c, column=9, value=fname)
            ws.cell(row=c, column=10, value=foccupation)
            ws.cell(row=c, column=11, value=institutename)
            ws.cell(row=c, column=12, value=course)
            ws.cell(row=c, column=13, value=presentyear)
            ws.cell(row=c, column=14, value=percentage12)
            ws.cell(row=c, column=15, value=cgpa)

            c = c+1
        # Save the file
    wb.save("media/xl_sheets1/all_xl_file1.xlsx")


# $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
#                           MAIL
# $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$


def send_email(to, name, mobile):
    name = name.upper()
    mobile = mobile
    message = EmailMessage()
    message['subject'] = "summer training-2020"
    message['from'] = "user email"
    message['to'] = to
    message.set_content("hello ")
    html_message = open("STPApp/templates/STPApp/mail.html",
                        encoding="utf8").read().replace("Taat", name+" - "+mobile)
    message.add_alternative(html_message, subtype='html')
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login("username", "password")
        smtp.send_message(message)


@login_required(login_url='officerlogin')
def sendmails(request):
    RegForms = RegForm.objects.all()
    total_no = len(RegForms)
    for i in range(total_no):
        if RegForms[i].counter == 2:
            name = RegForms[i].name
            mobile = RegForms[i].mobile
            email = RegForms[i].email
            send_email(email, name, mobile)
    messages.success(request, "Mail sent")
    return redirect("officer")


# ! Teachers Login
def handleteacherLogin(request):
    if request.method == 'POST':
        teacherusername = request.POST['teacherusername']
        teacherpass = request.POST['teacherpass']
        user = authenticate(username=teacherusername, password=teacherpass)
        if user is not None:
            usernameval1 = str(user.username)
            name = usernameval1[2:5]
            if name == "TEC":
                login(request, user)
                return redirect("/mentor")
            else:
                messages.error(request, "Invalid credentials")
                return redirect("/")
        else:
            messages.error(request, "Invalid credentials")
            return redirect("/")
    return HttpResponse('404 - Not found')


def handleteacherLogout(request):

    logout(request)
    messages.success(request, "Successfully logout")
    return redirect("/")


@login_required(login_url='teacherlogin')
def handleteacher(request):
    RegForms = RegForm.objects.all()
    n = len(RegForms)
    Tuploads = TUpload.objects.all()
    check = TUpload.objects.filter(tusername='17TEC0116')
    s1 = "Null"
    s2 = "Null"
    s3 = "NUll"
    if Tuploads[0].treview1:
        s1 = "Uploaded"
    if Tuploads[0].treview2:
        s2 = "Uploaded"
    if Tuploads[0].treview3:
        s3 = "Uploaded"

    params = {'allforms': RegForms, 'range': range(
        n), 'teacherUpload': Tuploads, 'status1': s1, 'status2': s2, 'status3': s3}
    return render(request, 'STPApp/teacher.html', params)


@login_required(login_url='teacherlogin')
def treview1test(request):
    if request.method == 'POST':
        filename1 = request.FILES['filename1']
        file1_name = request.FILES['filename1'].name
        fs = FileSystemStorage()
        file = fs.save(filename1.name, filename1)
        fileurl = fs.url(file)
        report = file1_name
        TUpload.objects.filter(
            tusername='17TEC0116').update(treview1=filename1)
        print("working")
    return redirect(handleteacher)


@login_required(login_url='teacherlogin')
def treview2test(request):
    if request.method == 'POST':
        filename2 = request.FILES['filename2']
        file2_name = request.FILES['filename2'].name
        fs = FileSystemStorage()
        file = fs.save(filename2.name, filename2)
        fileurl = fs.url(file)
        report = file2_name
        TUpload.objects.filter(
            tusername='17TEC0116').update(treview2=filename2)
        print("working")
    return redirect(handleteacher)


@login_required(login_url='teacherlogin')
def treview3test(request):
    if request.method == 'POST':
        filename3 = request.FILES['filename3']
        file3_name = request.FILES['filename3'].name
        fs = FileSystemStorage()
        file = fs.save(filename3.name, filename3)
        fileurl = fs.url(file)
        report = file3_name
        TUpload.objects.filter(
            tusername='17TEC0116').update(treview3=filename3)
        print("working")
    return redirect(handleteacher)


# ! Student Login
def handlestudentLogin(request):
    if request.method == 'POST':
        studentusername = request.POST['studentusername']
        studentpass = request.POST['studentpass']
        user = authenticate(username=studentusername, password=studentpass)
        if user is not None:
            usernameval1 = str(user.username)
            name = usernameval1[2:5]
            if name == "STU":
                login(request, user)
                return redirect("/studentAcc")
            else:
                messages.error(request, "Invalid credentials")
                return redirect("/")
        else:
            messages.error(request, "Invalid credentials")
            return redirect("/")
    return HttpResponse('404 - Not found')


def handlestudentLogout(request):

    logout(request)
    messages.success(request, "Successfully logout")
    return redirect("/")


@login_required(login_url='studentlogin')
def handlestudent(request):
    RegForms = RegForm.objects.all()
    tupload = TUpload.objects.all()
    tfileUp1 = tupload[0].treview1
    tfileUp2 = tupload[0].treview2
    tfileUp3 = tupload[0].treview3

    n = len(RegForms)
    print(request.user)
    params = {'allforms': RegForms, 'range': range(
        n), 'tfile1': tfileUp1, 'tfile2': tfileUp2, 'tfile3': tfileUp3}
    return render(request, 'STPApp/student.html', params)
